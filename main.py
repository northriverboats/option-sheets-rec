import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime

# Put your data at the top
pickles = [
	[r"K:\Links\2020\Options\options.pickle", r"output\Boat Options Parts Listing.xlsx"],
	[r"K:\Links\2020\Yamaha Rigging\yamaha rigging.pickle", r"output\Yamaha Options Parts Listing.xlsx"],
	[r"K:\Links\2020\Mercury Rigging\mercury rigging.pickle", r"output\Mercury Options Parts Listing.xlsx"],
	[r"K:\Links\2020\Honda Rigging\honda rigging.pickle", r"output\Honda Options Parts Listing.xlsx"],
]

RED_BOLD = Font(bold=True, underline="single", color="FF0000")
RED = Font(color="FF0000")
BLUE = Font(color="0000FF")


def load_pickle(file_name):
	with open(file_name, "rb") as file:
		return pickle.load(file)


def set_font_for_notes(option, ws, row):
	# if len(option["OUTFITTING PARTS"]) + len(option["CANVAS PARTS"]) + len(option["FABRICATION PARTS"]) + len(option["PAINT PARTS"]) > 0:
	if len(option["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  option["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = RED_BOLD
	if len(option["EOS OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  option["EOS OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = BLUE
	return row


def process_sections(option, ws, length, option_key, row):
	for section_name, section_options, count in [
		[
			'Paint', 
			option["PAINT PARTS"], 
			len(option["PAINT PARTS"]),
		],
		[
			"Outfutting", 
			option["OUTFITTING PARTS"] + option["CANVAS PARTS"], 
			len(option["OUTFITTING PARTS"]) + len(option["CANVAS PARTS"]),
		],
		[
			'Fabrication',
			option["FABRICATION PARTS"], 
			len(option["FABRICATION PARTS"])
		],
		[
			'Paint', 
			option["PAINT PARTS"],
			len(option["PAINT PARTS"]),
		],
	]:
		if count > 0:
			row += 1
			ws.cell(row = row, column = 1).value = option_key + " " + section_name
			ws.cell(row = row, column = 1).font = RED_BOLD
			ws.cell(row = row, column = 2).value = option["OPTION NAME"]
			ws.cell(row = row, column = 2).font = RED_BOLD

		if count > 0:
			for item in section_options:
				row += 1
				print(option_key, item["PART NUMBER"])
				ws.cell(row = row, column = 1).value = item["VENDOR"]
				ws.cell(row = row, column = 1).alignment = Alignment(horizontal='left')
				ws.cell(row = row, column = 2).value = item["VENDOR PART"]
				ws.cell(row = row, column = 2).alignment = Alignment(horizontal='left')
				ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
				ws.cell(row = row, column = 4).value = float(item["PRICE"])
				ws.cell(row = row, column = 4).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
				ws.cell(row = row, column = 5).value = item["UOM"]
				ws.cell(row = row, column = 6).value = item[length + " QTY"]
				ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
				ws.cell(row = row, column = 7).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
				ws.cell(row = row, column = 8).value = 0
				ws.cell(row = row, column = 8).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
				ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
				ws.cell(row = row, column = 9).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

		if len(option["PAINT PARTS"]) > 0:
				ws.cell(row = row, column = 1).value = option_key + " Paint"
		if len(option["PAINT PARTS"]) > 0:
			for item in option["PAINT PARTS"]:
				pass
	return row


def process_options(options, ws, length):
	row = 1
	for option_key in sorted(options):
		option = options[option_key]
		row = set_font_for_notes(option, ws, row)
		row = process_sections(option, ws, length, option_key, row)



def process_lengths(wb, options, file_name):
	lengths = [k.split(' ')[0] for k in options[next(iter(options))] if " TOTAL COST" in k]
	for length in lengths:
		ws = wb["L" + length]
		process_options(options, ws, length)
	wb.save(file_name)


def process_all_options(file_name_in, file_name_out):
	wb = load_workbook(r"templates\CostingSheetTemplate.xlsx")
	options =  load_pickle(file_name_in)
	process_lengths(wb, options, file_name_out)


for pickle_file, output in pickles:
	process_all_options(pickle_file, output)